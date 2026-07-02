"""Worker-side QA/Pro processors for CogniSweep async task handoffs.

The Streamlit app queues heavy workflows by storing uploaded files through the
object-storage adapter and sending a compact manifest to async_task_worker.py.
This module consumes that same manifest outside the UI, updates the durable task
lifecycle record, and writes downloadable result files.
"""
from __future__ import annotations

import argparse
import csv
import hashlib
import io
import json
import logging
import mimetypes
import os
import re
import tempfile
import time
import uuid
import zipfile
from datetime import datetime, timezone
from pathlib import Path
from typing import Any, Dict, Iterable, List, Optional, Tuple

import pandas as pd
import defusedxml.ElementTree as ET
import requests
from openpyxl import Workbook
from openpyxl.styles import Alignment, Font, PatternFill

from app_runtime_config import runtime_env
from cloud_object_storage import build_object_key, object_storage_provider, put_file, signed_url_for_key
from pro_reconstruction import build_export_source_asset_from_bytes, sentence_segment_rows_for_pro
from production_persistence import (
    fetch_saas_records,
    log_persistent_usage_event,
    save_persistent_editor_job,
    save_saas_record,
)

try:
    from translator_router import translate_batch
except ImportError as exc:  # pragma: no cover
    translate_batch = None
    logging.getLogger(__name__).warning("translator_router import failed: %s", exc)

LOGGER = logging.getLogger("errorsweep.async_workflow_processor")
SERVICE_NAME = "errorsweep-async-workflow-processor"
TERMINAL_STATUSES = {"completed", "failed", "cancelled", "needs_review"}
MAX_DOCX_BYTES = 50 * 1024 * 1024
MAX_DOCX_UNCOMPRESSED_BYTES = 150 * 1024 * 1024
MAX_DOCX_XML_BYTES = 25 * 1024 * 1024
MAX_MANIFEST_BYTES = 50 * 1024 * 1024
ALLOWED_MANIFEST_SUFFIXES = {
    ".csv",
    ".docx",
    ".json",
    ".md",
    ".srt",
    ".tsv",
    ".txt",
    ".vtt",
    ".xlsm",
    ".xlsx",
    ".zip",
}


def safe_text(value: Any) -> str:
    return "" if value is None else str(value).strip()


def safe_filename(value: Any, fallback: str = "file") -> str:
    cleaned = re.sub(r"[^A-Za-z0-9_. -]+", "_", safe_text(value)).strip(" .")
    return cleaned or fallback


def env_value(name: str, default: str = "") -> str:
    return runtime_env(name, default)


def env_int(name: str, default: int, minimum: int = 1) -> int:
    try:
        value = int(safe_text(env_value(name)))
    except (TypeError, ValueError):
        value = default
    return max(minimum, value)


def now_iso() -> str:
    return datetime.now(timezone.utc).isoformat()


def worker_root() -> Path:
    configured = safe_text(env_value("ERRORSWEEP_ASYNC_WORKER_DIR"))
    root = Path(configured) if configured else Path(tempfile.gettempdir()) / "errorsweep_async_worker"
    root.mkdir(parents=True, exist_ok=True)
    return root


def results_root() -> Path:
    configured = safe_text(env_value("ERRORSWEEP_ASYNC_RESULT_DIR"))
    root = Path(configured) if configured else worker_root() / "results"
    root.mkdir(parents=True, exist_ok=True)
    return root


def spool_path(task_id: str) -> Path:
    safe = "".join(ch for ch in safe_text(task_id) if ch.isalnum() or ch in {"-", "_"})[:96]
    return worker_root() / f"{safe or uuid.uuid4().hex}.json"


def read_spool(task_id: str) -> Optional[Dict[str, Any]]:
    path = spool_path(task_id)
    if not path.exists():
        return None
    try:
        data = json.loads(path.read_text(encoding="utf-8"))
        return data if isinstance(data, dict) else None
    except (OSError, json.JSONDecodeError, TypeError) as exc:
        LOGGER.warning("Unable to read worker spool %s: %s", path, exc)
        return None


def write_spool_status(task_id: str, payload: Dict[str, Any], status_record: Dict[str, Any]) -> None:
    path = spool_path(task_id)
    document = read_spool(task_id) or {"service": SERVICE_NAME, "task_id": task_id, "payload": payload}
    document["status_record"] = status_record
    document["last_processor_update_at"] = now_iso()
    tmp = path.with_suffix(".tmp")
    tmp.write_text(json.dumps(document, ensure_ascii=False, indent=2, default=safe_text), encoding="utf-8")
    os.replace(tmp, path)


def task_id_from_payload(payload: Dict[str, Any]) -> str:
    return safe_text(payload.get("task_id") or payload.get("id") or uuid.uuid4().hex)


def persisted_task(task_id: str) -> Dict[str, Any]:
    try:
        for item in fetch_saas_records("task_queue", limit=1000, include_all_workspaces=True):
            if safe_text(item.get("id")) == safe_text(task_id):
                return dict(item)
    except Exception as exc:
        LOGGER.warning("Unable to fetch task %s: %s", task_id, exc)
    return {"id": task_id, "created_at": now_iso()}


def update_task(
    task_id: str,
    payload: Dict[str, Any],
    *,
    status: str,
    progress: int,
    processed_units: int = 0,
    total_units: int = 0,
    result_ref: str = "",
    error: str = "",
    metadata: Optional[Dict[str, Any]] = None,
) -> Dict[str, Any]:
    existing = persisted_task(task_id)
    base_metadata = existing.get("metadata_json") if isinstance(existing.get("metadata_json"), dict) else {}
    top_metadata = payload.get("metadata") if isinstance(payload.get("metadata"), dict) else {}
    workflow_payload = payload.get("payload") if isinstance(payload.get("payload"), dict) else {}
    workspace = safe_text(payload.get("workspace") or existing.get("workspace") or top_metadata.get("workspace") or "Unassigned")
    user_email = safe_text(payload.get("user_email") or existing.get("user_email") or top_metadata.get("user_email") or "async-worker@cognisweep.local")
    now = now_iso()
    record = {
        **existing,
        "id": task_id,
        "workspace": workspace,
        "user_email": user_email,
        "task_type": safe_text(payload.get("task_type") or existing.get("task_type") or workflow_payload.get("workflow") or "external_task"),
        "label": safe_text(payload.get("label") or existing.get("label") or top_metadata.get("file_name") or workflow_payload.get("workflow") or task_id),
        "status": status,
        "progress": max(0, min(100, int(progress or 0))),
        "processed_units": int(processed_units or existing.get("processed_units") or 0),
        "total_units": int(total_units or existing.get("total_units") or 0),
        "result_ref": safe_text(result_ref or existing.get("result_ref") or f"{SERVICE_NAME}:{task_id}"),
        "error": safe_text(error)[:1000],
        "metadata_json": {
            **base_metadata,
            **top_metadata,
            **(metadata or {}),
            "processor_attached": True,
            "processor": SERVICE_NAME,
            "processor_status": status,
            "last_processor_update_at": now,
        },
        "updated_at": now,
    }
    if status == "running" and not safe_text(record.get("started_at")):
        record["started_at"] = now
    if status in TERMINAL_STATUSES and not safe_text(record.get("finished_at")):
        record["finished_at"] = now
    persisted = save_saas_record("task_queue", record, user={"email": user_email, "workspace": workspace})
    write_spool_status(task_id, payload, persisted)
    return persisted


def manifest_file_name(manifest: Dict[str, Any]) -> str:
    return safe_text(manifest.get("file_name") or Path(safe_text(manifest.get("storage_key"))).name or "uploaded_file")


def manifest_max_bytes() -> int:
    return env_int("ERRORSWEEP_ASYNC_MAX_MANIFEST_BYTES", MAX_MANIFEST_BYTES, minimum=1024)


def validate_manifest_for_read(manifest: Dict[str, Any]) -> None:
    if not isinstance(manifest, dict):
        raise ValueError("Queued file manifest is invalid.")
    file_name = manifest_file_name(manifest)
    suffix = Path(file_name).suffix.lower()
    if suffix not in ALLOWED_MANIFEST_SUFFIXES:
        raise ValueError(f"Queued file manifest uses an unsupported file type: {suffix or 'missing extension'}.")
    declared_size = safe_text(manifest.get("size_bytes") or manifest.get("content_length") or manifest.get("bytes"))
    if declared_size:
        try:
            parsed_size = int(float(declared_size))
        except (TypeError, ValueError):
            LOGGER.warning("Queued file manifest has invalid size metadata for %s: %s", file_name, declared_size)
        else:
            if parsed_size > manifest_max_bytes():
                raise ValueError("Queued file manifest exceeds async worker file limits.")


def checked_local_manifest_bytes(path: Path, file_name: str) -> bytes:
    max_bytes = manifest_max_bytes()
    if path.stat().st_size > max_bytes:
        raise ValueError(f"Queued file {file_name} exceeds async worker file limits.")
    data = path.read_bytes()
    if len(data) > max_bytes:
        raise ValueError(f"Queued file {file_name} exceeds async worker file limits.")
    return data


def checked_remote_manifest_bytes(url: str, file_name: str) -> bytes:
    max_bytes = manifest_max_bytes()
    timeout = int(env_value("ERRORSWEEP_ASYNC_FILE_TIMEOUT", "120"))
    response = requests.get(url, timeout=timeout, stream=True)
    try:
        response.raise_for_status()
        declared_length = safe_text(response.headers.get("Content-Length") if hasattr(response, "headers") else "")
        if declared_length:
            try:
                parsed_length = int(declared_length)
            except (TypeError, ValueError):
                LOGGER.warning("Queued file %s has invalid Content-Length: %s", file_name, declared_length)
            else:
                if parsed_length > max_bytes:
                    raise ValueError(f"Queued file {file_name} exceeds async worker file limits.")
        chunks: List[bytes] = []
        total = 0
        for chunk in response.iter_content(chunk_size=1024 * 1024):
            if not chunk:
                continue
            total += len(chunk)
            if total > max_bytes:
                raise ValueError(f"Queued file {file_name} exceeds async worker file limits.")
            chunks.append(chunk)
        return b"".join(chunks)
    finally:
        response.close()


def manifest_bytes(manifest: Dict[str, Any]) -> bytes:
    """Read a queued input/rules file from local fallback, public URL, or signed URL."""
    validate_manifest_for_read(manifest)
    file_name = manifest_file_name(manifest)
    local_candidates = [
        safe_text(manifest.get("local_path")),
        safe_text(manifest.get("storage_key")) if safe_text(manifest.get("storage_provider")).lower() in {"", "local"} else "",
    ]
    for candidate in local_candidates:
        if candidate and Path(candidate).exists():
            return checked_local_manifest_bytes(Path(candidate), file_name)

    public_url = safe_text(manifest.get("public_url"))
    if public_url:
        return checked_remote_manifest_bytes(public_url, file_name)

    storage_key = safe_text(manifest.get("storage_key"))
    if storage_key:
        url = signed_url_for_key(storage_key, expires_in=900)
        if url and url != storage_key:
            return checked_remote_manifest_bytes(url, file_name)
    raise FileNotFoundError(f"Unable to read queued file manifest: {manifest_file_name(manifest)}")


def normalize_header(value: Any) -> str:
    return re.sub(r"[^a-z0-9]+", "", safe_text(value).lower())


SOURCE_KEYS = {
    "source", "sourcetext", "sourcecopy", "original", "originaltext", "english", "en",
    "string", "stringidtext", "text", "sourcecontent",
}
TARGET_KEYS = {
    "target", "targettext", "translation", "translated", "translatedtext", "localized",
    "localizedtext", "targetcopy", "reviewedtarget", "reviewedtranslation", "final",
}
TARGET_BLOCK_KEYS = {"targetlanguage", "targetlang", "language", "locale", "targetlocale"}
ID_KEYS = {"id", "segmentid", "key", "stringid", "identifier", "segment"}


def choose_columns(columns: Iterable[Any]) -> Tuple[Optional[str], Optional[str], Optional[str]]:
    normalized = {str(col): normalize_header(col) for col in columns}
    source_col = next((col for col, key in normalized.items() if key in SOURCE_KEYS), None)
    target_col = next((col for col, key in normalized.items() if key in TARGET_KEYS and key not in TARGET_BLOCK_KEYS), None)
    id_col = next((col for col, key in normalized.items() if key in ID_KEYS), None)
    if not source_col or not target_col:
        textish = [str(col) for col in columns if normalized.get(str(col)) not in TARGET_BLOCK_KEYS]
        if not source_col and textish:
            source_col = textish[0]
        if not target_col and len(textish) > 1:
            target_col = next((col for col in textish[1:] if col != source_col), None)
    return source_col, target_col, id_col


def dataframe_to_rows(df: pd.DataFrame, sheet: str = "") -> List[Dict[str, Any]]:
    if df.empty:
        return []
    df = df.fillna("")
    source_col, target_col, id_col = choose_columns(df.columns)
    rows: List[Dict[str, Any]] = []
    if not source_col and not target_col:
        return rows
    for idx, record in enumerate(df.to_dict(orient="records"), start=1):
        source = safe_text(record.get(source_col, "")) if source_col else ""
        target = safe_text(record.get(target_col, "")) if target_col else ""
        if not source and not target:
            continue
        rows.append({
            "id": safe_text(record.get(id_col, "")) if id_col else str(len(rows) + 1),
            "location": f"{sheet} Row {idx}".strip(),
            "source": source,
            "target": target,
            "status": "Existing" if target else "Untranslated",
            "match": "Existing" if target else "Untranslated",
        })
    return rows


def parse_srt_or_vtt(text: str) -> List[Dict[str, Any]]:
    blocks = re.split(r"\n\s*\n", text.replace("\r\n", "\n").strip())
    rows: List[Dict[str, Any]] = []
    for block in blocks:
        lines = [line.strip() for line in block.splitlines() if line.strip()]
        if not lines:
            continue
        if lines and re.fullmatch(r"\d+", lines[0]):
            lines = lines[1:]
        timing = lines[0] if lines and "-->" in lines[0] else ""
        body = " ".join(lines[1:] if timing else lines)
        if body:
            rows.append({"id": str(len(rows) + 1), "location": timing or f"Segment {len(rows) + 1}", "source": body, "target": "", "status": "Untranslated", "match": "Untranslated"})
    return rows


def safe_docx_document_xml(data: bytes) -> bytes:
    max_bytes = env_int("ERRORSWEEP_ASYNC_MAX_DOCX_BYTES", MAX_DOCX_BYTES)
    max_uncompressed = env_int("ERRORSWEEP_ASYNC_MAX_DOCX_UNCOMPRESSED_BYTES", MAX_DOCX_UNCOMPRESSED_BYTES)
    max_xml = env_int("ERRORSWEEP_ASYNC_MAX_DOCX_XML_BYTES", MAX_DOCX_XML_BYTES)
    if len(data or b"") > max_bytes:
        raise ValueError("DOCX file is too large for async worker parsing.")
    with zipfile.ZipFile(io.BytesIO(data)) as archive:
        total_uncompressed = sum(max(0, info.file_size) for info in archive.infolist())
        if total_uncompressed > max_uncompressed:
            raise ValueError("DOCX expanded size exceeds async worker limits.")
        try:
            info = archive.getinfo("word/document.xml")
        except KeyError as exc:
            raise ValueError("DOCX is missing word/document.xml.") from exc
        if info.file_size > max_xml:
            raise ValueError("DOCX document XML exceeds async worker limits.")
        document_xml = archive.read(info)
    upper = document_xml.upper()
    if b"<!DOCTYPE" in upper or b"<!ENTITY" in upper:
        raise ValueError("DOCX XML contains prohibited DTD/entity declarations.")
    return document_xml


def docx_text_from_element(element: Any, ns: Dict[str, str]) -> str:
    pieces: List[str] = []
    for node in element.iter():
        if node.tag == f"{{{ns['w']}}}t" and node.text:
            pieces.append(node.text)
        elif node.tag == f"{{{ns['w']}}}tab":
            pieces.append("\t")
        elif node.tag in {f"{{{ns['w']}}}br", f"{{{ns['w']}}}cr"}:
            pieces.append("\n")
    return safe_text("".join(pieces))


def parse_docx(data: bytes) -> List[Dict[str, Any]]:
    document_xml = safe_docx_document_xml(data)
    root = ET.fromstring(document_xml, forbid_dtd=True, forbid_entities=True)
    ns = {"w": "http://schemas.openxmlformats.org/wordprocessingml/2006/main"}
    rows: List[Dict[str, Any]] = []
    for table_idx, table in enumerate(root.findall(".//w:tbl", ns), start=1):
        for row_idx, table_row in enumerate(table.findall("./w:tr", ns), start=1):
            cells = [docx_text_from_element(cell, ns) for cell in table_row.findall("./w:tc", ns)]
            filled = [cell for cell in cells if cell]
            if not filled:
                continue
            rows.append({
                "id": str(len(rows) + 1),
                "location": f"Table {table_idx} Row {row_idx}",
                "source": filled[0],
                "target": filled[1] if len(filled) > 1 else "",
                "status": "Existing" if len(filled) > 1 else "Untranslated",
                "match": "Existing" if len(filled) > 1 else "Untranslated",
            })
    for para_idx, para in enumerate(root.findall("./w:body/w:p", ns), start=1):
        text = docx_text_from_element(para, ns)
        if text:
            rows.append({"id": str(len(rows) + 1), "location": f"Paragraph {para_idx}", "source": text, "target": "", "status": "Untranslated", "match": "Untranslated"})
    return rows


def parse_json_rows(data: bytes) -> List[Dict[str, Any]]:
    parsed = json.loads(data.decode("utf-8-sig"))
    if isinstance(parsed, dict):
        items = parsed.get("items") or parsed.get("segments") or parsed.get("rows") or list(parsed.values())
    else:
        items = parsed
    rows: List[Dict[str, Any]] = []
    if not isinstance(items, list):
        return rows
    for idx, item in enumerate(items, start=1):
        if isinstance(item, dict):
            source_col, target_col, id_col = choose_columns(item.keys())
            source = safe_text(item.get(source_col, "")) if source_col else ""
            target = safe_text(item.get(target_col, "")) if target_col else ""
            row_id = safe_text(item.get(id_col, "")) if id_col else str(idx)
        else:
            source, target, row_id = safe_text(item), "", str(idx)
        if source or target:
            rows.append({"id": row_id or str(idx), "location": f"JSON row {idx}", "source": source, "target": target, "status": "Existing" if target else "Untranslated", "match": "Existing" if target else "Untranslated"})
    return rows


def parse_upload_rows(data: bytes, file_name: str) -> List[Dict[str, Any]]:
    suffix = Path(file_name).suffix.lower()
    if suffix in {".xlsx", ".xlsm"}:
        sheets = pd.read_excel(io.BytesIO(data), sheet_name=None, dtype=str, engine="openpyxl")
        rows: List[Dict[str, Any]] = []
        for sheet_name, frame in sheets.items():
            rows.extend(dataframe_to_rows(frame, sheet=safe_text(sheet_name)))
        return rows
    if suffix == ".csv":
        return dataframe_to_rows(pd.read_csv(io.BytesIO(data), dtype=str, keep_default_na=False), sheet="CSV")
    if suffix == ".tsv":
        return dataframe_to_rows(pd.read_csv(io.BytesIO(data), sep="\t", dtype=str, keep_default_na=False), sheet="TSV")
    if suffix == ".docx":
        return parse_docx(data)
    if suffix in {".srt", ".vtt"}:
        return parse_srt_or_vtt(data.decode("utf-8-sig", errors="replace"))
    if suffix == ".json":
        return parse_json_rows(data)
    text = data.decode("utf-8-sig", errors="replace")
    rows = []
    for idx, line in enumerate(text.splitlines(), start=1):
        line = safe_text(line)
        if line:
            rows.append({"id": str(len(rows) + 1), "location": f"Line {idx}", "source": line, "target": "", "status": "Untranslated", "match": "Untranslated"})
    return rows


def parse_rules_zip_from_bytes(data: bytes) -> Dict[str, Any]:
    rules = {"glossary": [], "dnt": [], "instructions": []}
    if not data:
        return rules
    try:
        with zipfile.ZipFile(io.BytesIO(data)) as zf:
            for info in zf.infolist()[:200]:
                if info.is_dir() or info.file_size > 2 * 1024 * 1024:
                    continue
                name = info.filename.lower()
                raw = zf.read(info).decode("utf-8-sig", errors="replace")
                if "dnt" in name or "do-not-translate" in name:
                    for line in raw.splitlines():
                        term = safe_text(re.split(r"[,;\t]", line)[0])
                        if term:
                            rules["dnt"].append({"term": term})
                elif "gloss" in name or "term" in name:
                    reader = csv.reader(io.StringIO(raw))
                    for row in reader:
                        if len(row) >= 2 and safe_text(row[0]) and safe_text(row[1]):
                            rules["glossary"].append({"source": safe_text(row[0]), "target": safe_text(row[1])})
                elif name.endswith((".txt", ".md")):
                    for line in raw.splitlines():
                        line = safe_text(line)
                        if line:
                            rules["instructions"].append(line)
    except (OSError, UnicodeDecodeError, zipfile.BadZipFile, zipfile.LargeZipFile, csv.Error, ValueError) as exc:
        rules["warnings"] = [f"Rules ZIP could not be parsed by worker: {exc}"]
    return rules


def merge_text_rules(rules: Dict[str, Any], text: str, source_name: str = "QA context") -> None:
    """Extract deterministic glossary/DNT hints from inline QA instructions or references."""
    value = safe_text(text)
    if not value:
        return
    rules.setdefault("instructions", []).append({"text": value[:5000], "source": source_name})
    for raw_line in value.splitlines():
        line = safe_text(raw_line)
        if not line or len(line) > 700:
            continue
        dnt_match = re.match(r"^(?:dnt|do\s*not\s*translate|locked\s*term|keep(?:\s+unchanged)?)\s*[:=-]\s*(.+)$", line, flags=re.I)
        if dnt_match:
            for term in re.split(r"[;,]|\t|\|", dnt_match.group(1)):
                term = safe_text(term).strip("-* ")
                if term:
                    rules.setdefault("dnt", []).append({"term": term, "source": source_name})
            continue
        candidate = re.sub(r"^(?:glossary|term|terminology)\s*[:=-]\s*", "", line, flags=re.I).strip()
        if "->" in candidate or "=>" in candidate:
            sep = "->" if "->" in candidate else "=>"
            source, target = [safe_text(part) for part in candidate.split(sep, 1)]
            source = re.sub(r"^(?:source|src|english|term)\s*[:=]\s*", "", source, flags=re.I).strip()
            target = re.sub(r"^(?:target|tgt|translation|preferred)\s*[:=]\s*", "", target, flags=re.I).strip()
            if source and target and source.lower() != target.lower():
                rules.setdefault("glossary", []).append({"source": source, "target": target, "source_name": source_name})
            elif source:
                rules.setdefault("dnt", []).append({"term": source, "source": source_name})


def translation_context_metadata(
    rules: Dict[str, Any],
    reference_context: Dict[str, Any],
    source_texts: List[str],
    target_language: str,
    domain: str,
) -> Dict[str, Any]:
    """Carry client context to async translation adapters."""
    rule_pack = rules if isinstance(rules, dict) else {}
    glossary: List[Dict[str, str]] = []
    for item in rule_pack.get("glossary", []) or []:
        if not isinstance(item, dict):
            continue
        source = safe_text(item.get("source") or item.get("source_term"))
        target = safe_text(item.get("target") or item.get("target_term"))
        if source or target:
            glossary.append(
                {
                    "source": source,
                    "target": target,
                    "source_name": safe_text(item.get("source_name") or item.get("source")),
                }
            )
    dnt_terms = [
        safe_text(item.get("term") if isinstance(item, dict) else item)
        for item in rule_pack.get("dnt", []) or []
        if safe_text(item.get("term") if isinstance(item, dict) else item)
    ]
    instructions = [
        safe_text(item.get("text") if isinstance(item, dict) else item)
        for item in rule_pack.get("instructions", []) or []
        if safe_text(item.get("text") if isinstance(item, dict) else item)
    ]
    ref = reference_context if isinstance(reference_context, dict) else {}
    return {
        "target_language": safe_text(target_language),
        "domain": safe_text(domain),
        "dnt_terms": dnt_terms[:200],
        "glossary": glossary[:200],
        "instructions": instructions[:80],
        "reference_context": {
            "fileName": safe_text(ref.get("fileName")),
            "text": safe_text(ref.get("text"))[:12000],
        },
        "segment_count": len(source_texts or []),
        "sample_sources": [safe_text(text)[:500] for text in (source_texts or [])[:12] if safe_text(text)],
    }


def extract_placeholders(text: str) -> List[str]:
    return re.findall(r"\{\{[^{}]+\}\}|\{[^{}]+\}|%[0-9$.\-+]*[sdif]|</?[\w:-]+[^>]*>|https?://\S+|www\.\S+", text or "")


def extract_numbers(text: str) -> List[str]:
    return re.findall(r"(?<![\w.])-?\d+(?:[.,]\d+)*(?:%|[A-Za-z]{1,4})?", text or "")


def extract_emojis(text: str) -> List[str]:
    return re.findall(r"[\U0001F1E6-\U0001FAFF\u2600-\u27BF]", text or "")


def finding(row: Dict[str, Any], severity: str, category: str, issue: str, suggestion: str = "", rule_id: str = "") -> Dict[str, Any]:
    return {
        "Finding ID": f"QA-{uuid.uuid4().hex[:8].upper()}",
        "Segment ID": safe_text(row.get("id")),
        "Location": safe_text(row.get("location") or f"Segment {row.get('id', '')}"),
        "Source Text": safe_text(row.get("source")),
        "Target Text": safe_text(row.get("target")),
        "Suggested Target": safe_text(suggestion),
        "Error Category": safe_text(category),
        "Severity": safe_text(severity),
        "Issue": safe_text(issue),
        "Explanation": safe_text(issue),
        "Check": "CogniSweep worker deterministic QA",
        "Confidence": "High",
        "Rule ID": safe_text(rule_id or f"worker.{normalize_header(category) or 'qa'}"),
    }


def qa_findings_for_row(row: Dict[str, Any], rules: Dict[str, Any]) -> List[Dict[str, Any]]:
    source = safe_text(row.get("source"))
    target = safe_text(row.get("target"))
    out: List[Dict[str, Any]] = []
    if source and not target:
        out.append(finding(row, "Major", "Completeness", "Target translation is blank.", "Add the missing translation."))
    if source and target and source == target:
        out.append(finding(row, "Critical", "Untranslated Text", "Target is identical to source.", "Translate the source text or confirm it is intentionally unchanged."))
    for ph in extract_placeholders(source):
        if ph and ph not in target:
            out.append(finding(row, "Major", "Placeholder", f"Missing protected placeholder/tag: {ph}", f"Keep {ph} unchanged in the target."))
    for emoji in sorted(set(extract_emojis(source)) - set(extract_emojis(target))):
        out.append(finding(row, "Major", "Emoji/Icon", f"Source emoji/icon is missing in target: {emoji}", f"Keep {emoji} from the source unless the client rules say otherwise."))
    source_numbers = extract_numbers(source)
    target_numbers = extract_numbers(target)
    for number in source_numbers:
        if number and number not in target_numbers and number not in target:
            out.append(finding(row, "Major", "Number", f"Source number is missing or changed: {number}", f"Keep or correctly localize {number}."))
    if "\u200c" in source and "\u200c" not in target:
        out.append(finding(row, "Major", "Zero Width Non-Joiner", "Source uses Zero Width Non-Joiner but target does not.", "Preserve required ZWNJ usage for the target language."))
    for item in rules.get("dnt", []) or []:
        term = safe_text(item.get("term") if isinstance(item, dict) else item)
        if term and re.search(re.escape(term), source, flags=re.I) and not re.search(re.escape(term), target, flags=re.I):
            out.append(finding(row, "Major", "Do Not Translate", f"DNT term is missing or changed: {term}", f"Keep {term} unchanged."))
    for item in rules.get("glossary", []) or []:
        if not isinstance(item, dict):
            continue
        source_term = safe_text(item.get("source"))
        target_term = safe_text(item.get("target"))
        if source_term and target_term and re.search(re.escape(source_term), source, flags=re.I) and target_term.lower() not in target.lower():
            out.append(finding(row, "Minor", "Terminology", f"Glossary target term is missing: {target_term}", f"Use glossary term: {target_term}"))
    return out


def qa_status(row_findings: List[Dict[str, Any]]) -> str:
    severities = {safe_text(item.get("Severity")).lower() for item in row_findings}
    if "critical" in severities:
        return "Fail"
    if "major" in severities:
        return "Needs Review"
    if row_findings:
        return "Review"
    return "Pass"


def build_qa_result(rows: List[Dict[str, Any]], rules: Dict[str, Any]) -> Tuple[List[Dict[str, Any]], List[Dict[str, Any]], Dict[str, Any]]:
    report_rows: List[Dict[str, Any]] = []
    segment_rows: List[Dict[str, Any]] = []
    for row in rows:
        row_findings = qa_findings_for_row(row, rules)
        report_rows.extend(row_findings)
        segment_rows.append({
            **row,
            "status": qa_status(row_findings),
            "issues": "; ".join(f"{item['Severity']}: {item['Error Category']}" for item in row_findings[:4]),
        })
    penalty = sum(10 if item.get("Severity") == "Critical" else 5 if item.get("Severity") == "Major" else 1 for item in report_rows)
    score = max(0, round(100 - (penalty / max(len(rows), 1)), 2))
    summary = {
        "total_segments": len(rows),
        "total_findings": len(report_rows),
        "qa_score": score,
        "result": "Pass" if score >= 95 else "Fail",
        "review_segments": sum(1 for row in segment_rows if safe_text(row.get("status")) != "Pass"),
    }
    return segment_rows, report_rows, summary


def auto_fit(ws: Any) -> None:
    for column_cells in ws.columns:
        letter = column_cells[0].column_letter
        max_len = min(70, max(len(safe_text(cell.value)) for cell in column_cells) + 2)
        ws.column_dimensions[letter].width = max(12, max_len)


def create_qa_workbook(segment_rows: List[Dict[str, Any]], findings: List[Dict[str, Any]], summary: Dict[str, Any]) -> bytes:
    wb = Workbook()
    ws = wb.active
    ws.title = "Summary"
    ws.append(["Metric", "Value"])
    for key in ["result", "qa_score", "total_segments", "total_findings", "review_segments"]:
        ws.append([key.replace("_", " ").title(), summary.get(key)])
    ws["A1"].font = ws["B1"].font = Font(bold=True, color="FFFFFF")
    ws["A1"].fill = ws["B1"].fill = PatternFill("solid", fgColor="111827")
    auto_fit(ws)

    ws_findings = wb.create_sheet("QA Findings")
    finding_columns = [
        "Finding ID", "Segment ID", "Source Text", "Target Text", "Suggested Target",
        "Error Category", "Severity", "Issue", "Explanation", "Check", "Confidence", "Rule ID",
    ]
    ws_findings.append(finding_columns)
    for item in findings:
        ws_findings.append([item.get(col, "") for col in finding_columns])
    for cell in ws_findings[1]:
        cell.font = Font(bold=True, color="FFFFFF")
        cell.fill = PatternFill("solid", fgColor="1F2937")
    auto_fit(ws_findings)

    ws_segments = wb.create_sheet("Segment Overview")
    segment_columns = ["id", "location", "source", "target", "status", "issues", "match"]
    ws_segments.append([col.replace("_", " ").title() for col in segment_columns])
    for row in segment_rows:
        ws_segments.append([row.get(col, "") for col in segment_columns])
    for cell in ws_segments[1]:
        cell.font = Font(bold=True, color="FFFFFF")
        cell.fill = PatternFill("solid", fgColor="1F2937")
    for sheet in wb.worksheets:
        for row in sheet.iter_rows():
            for cell in row:
                cell.alignment = Alignment(vertical="top", wrap_text=True)
    auto_fit(ws_segments)

    ws_notes = wb.create_sheet("Review Notes")
    ws_notes.append(["Note"])
    ws_notes.append(["Generated by CogniSweep async workflow processor. Review high/major findings before delivery."])
    auto_fit(ws_notes)

    buffer = io.BytesIO()
    wb.save(buffer)
    return buffer.getvalue()


def create_review_workbook(rows: List[Dict[str, Any]], summary: Dict[str, Any]) -> bytes:
    wb = Workbook()
    ws = wb.active
    ws.title = "Human Review"
    columns = ["id", "location", "source", "target", "status", "match", "qa_summary"]
    ws.append([col.replace("_", " ").title() for col in columns])
    for row in rows:
        ws.append([row.get(col, "") for col in columns])
    for cell in ws[1]:
        cell.font = Font(bold=True, color="FFFFFF")
        cell.fill = PatternFill("solid", fgColor="1F2937")
    auto_fit(ws)
    ws_summary = wb.create_sheet("Summary")
    ws_summary.append(["Metric", "Value"])
    for key, value in summary.items():
        ws_summary.append([safe_text(key).replace("_", " ").title(), value])
    auto_fit(ws_summary)
    buffer = io.BytesIO()
    wb.save(buffer)
    return buffer.getvalue()


def format_timed(seconds: Any, comma: bool = True) -> str:
    try:
        value = max(float(seconds), 0.0)
    except Exception:
        value = 0.0
    hours = int(value // 3600)
    minutes = int((value % 3600) // 60)
    secs = value % 60
    separator = "," if comma else "."
    return f"{hours:02d}:{minutes:02d}:{int(secs):02d}{separator}{int((secs % 1) * 1000):03d}"


def rows_to_csv_bytes(rows: List[Dict[str, Any]]) -> bytes:
    if not rows:
        return b""
    return pd.DataFrame(rows).to_csv(index=False).encode("utf-8-sig")


def rows_to_timed_text(rows: List[Dict[str, Any]], use_target: bool = True, vtt: bool = False) -> bytes:
    output: List[str] = ["WEBVTT", ""] if vtt else []
    for idx, row in enumerate(rows or [], start=1):
        text = safe_text(row.get("target" if use_target else "source")) or safe_text(row.get("source"))
        start = row.get("start", (idx - 1) * 4)
        end = row.get("end", (idx - 1) * 4 + 3)
        if not vtt:
            output.append(str(idx))
        output.append(f"{format_timed(start, comma=not vtt)} --> {format_timed(end, comma=not vtt)}")
        output.append(text)
        output.append("")
    return "\n".join(output).encode("utf-8")


def translated_rows_text(rows: List[Dict[str, Any]]) -> bytes:
    return "\n".join(
        safe_text(row.get("target")) or safe_text(row.get("source"))
        for row in rows or []
    ).encode("utf-8")


def create_translation_delivery_zip(
    rows: List[Dict[str, Any]],
    primary_manifest: Dict[str, Any],
    review_workbook: bytes = b"",
) -> bytes:
    source_name = safe_filename(manifest_file_name(primary_manifest) or "source_file")
    suffix = Path(source_name).suffix.lower()
    stem = Path(source_name).stem or "translated_file"
    if suffix == ".srt":
        translated_name = f"{stem}_translated.srt"
        translated_bytes = rows_to_timed_text(rows, use_target=True, vtt=False)
    elif suffix == ".vtt":
        translated_name = f"{stem}_translated.vtt"
        translated_bytes = rows_to_timed_text(rows, use_target=True, vtt=True)
    elif suffix == ".csv":
        translated_name = f"{stem}_translated.csv"
        translated_bytes = rows_to_csv_bytes(rows)
    elif suffix == ".txt":
        translated_name = f"{stem}_translated.txt"
        translated_bytes = translated_rows_text(rows)
    else:
        translated_name = f"{stem}_translated.txt"
        translated_bytes = translated_rows_text(rows)

    buffer = io.BytesIO()
    with zipfile.ZipFile(buffer, "w", compression=zipfile.ZIP_DEFLATED) as package:
        try:
            package.writestr(f"source_file/{source_name}", manifest_bytes(primary_manifest))
        except Exception as exc:
            LOGGER.warning("Unable to include source file in Pro delivery ZIP: %s", exc)
        package.writestr(f"translated_file/{translated_name}", translated_bytes)
        package.writestr("translated_segments.csv", rows_to_csv_bytes(rows))
        if review_workbook:
            package.writestr("human_review_workbook.xlsx", review_workbook)
    return buffer.getvalue()


def persist_result_file(data: bytes, file_name: str, workspace: str, user_email: str, purpose: str, object_id: str, mime_type: str) -> Dict[str, Any]:
    output_path = results_root() / f"{object_id}_{file_name}"
    output_path.write_bytes(data)
    sha = hashlib.sha256(data).hexdigest()
    key = build_object_key(workspace, purpose, object_id, file_name)
    storage = put_file(output_path, key, content_type=mime_type)
    manifest = save_saas_record(
        "files",
        {
            "workspace": workspace,
            "user_email": user_email,
            "file_name": file_name,
            "purpose": purpose,
            "mime_type": mime_type,
            "size_bytes": len(data),
            "sha256": sha,
            "storage_key": storage.get("storage_key", str(output_path)),
            "storage_provider": storage.get("storage_provider", object_storage_provider()),
            "storage_bucket": storage.get("storage_bucket", ""),
            "public_url": storage.get("public_url", ""),
            "local_path": storage.get("local_path", str(output_path)),
            "status": storage.get("status", "stored"),
            "created_at": now_iso(),
            "updated_at": now_iso(),
        },
        user={"email": user_email, "workspace": workspace},
    )
    return manifest


def load_task_inputs(payload: Dict[str, Any]) -> Tuple[List[Dict[str, Any]], Dict[str, Any], Dict[str, Any]]:
    task_payload = payload.get("payload") if isinstance(payload.get("payload"), dict) else {}
    input_files = task_payload.get("input_files") if isinstance(task_payload.get("input_files"), list) else []
    rules_files = task_payload.get("rules_files") if isinstance(task_payload.get("rules_files"), list) else []
    params = task_payload.get("parameters") if isinstance(task_payload.get("parameters"), dict) else {}
    if not input_files:
        raise ValueError("Queued workflow has no input file manifest.")
    primary = input_files[0]
    rows = parse_upload_rows(manifest_bytes(primary), manifest_file_name(primary))
    rules: Dict[str, Any] = {"glossary": [], "dnt": [], "instructions": []}
    for manifest in rules_files:
        parsed = parse_rules_zip_from_bytes(manifest_bytes(manifest))
        for key in ("glossary", "dnt", "instructions", "warnings"):
            rules.setdefault(key, [])
            rules[key].extend(parsed.get(key, []) or [])
    merge_text_rules(rules, safe_text(params.get("qa_inline_instructions")), "QA run instructions")
    reference_context = params.get("reference_context") if isinstance(params.get("reference_context"), dict) else {}
    merge_text_rules(
        rules,
        safe_text(reference_context.get("text")),
        safe_text(reference_context.get("fileName") or "Reference document"),
    )
    return rows, rules, primary


def process_qa_task(task_id: str, payload: Dict[str, Any]) -> Dict[str, Any]:
    workspace = safe_text(payload.get("workspace") or "Unassigned")
    user_email = safe_text(payload.get("user_email") or "async-worker@cognisweep.local")
    update_task(task_id, payload, status="running", progress=15)
    rows, rules, primary = load_task_inputs(payload)
    if not rows:
        raise ValueError("No QA-ready rows were detected in the queued input file.")
    update_task(task_id, payload, status="running", progress=45, total_units=len(rows), processed_units=0)
    segment_rows, findings, summary = build_qa_result(rows, rules)
    report = create_qa_workbook(segment_rows, findings, summary)
    report_manifest = persist_result_file(
        report,
        "CogniSweep_QA_Report.xlsx",
        workspace,
        user_email,
        "qa_report",
        task_id,
        "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
    )
    job = save_saas_record(
        "jobs",
        {
            "workspace": workspace,
            "user_email": user_email,
            "created": now_iso(),
            "type": "QA",
            "language": "",
            "status": "Completed",
            "segments": len(segment_rows),
            "note": f"Async QA report generated for {manifest_file_name(primary)}",
            "attachment_count": 1,
            "attachments_json": [report_manifest],
        },
        user={"email": user_email, "workspace": workspace},
    )
    log_persistent_usage_event(
        {"provider": "errorsweep_async_qa", "model": "deterministic_worker", "segments": len(segment_rows), "characters": sum(len(r.get("source", "")) + len(r.get("target", "")) for r in segment_rows), "requests": 1, "success": True},
        purpose="qa_workflow",
        segment_count=len(segment_rows),
        user={"email": user_email, "workspace": workspace},
        metadata={"task_id": task_id, "job_id": job.get("id"), "report_file_id": report_manifest.get("id")},
    )
    return update_task(
        task_id,
        payload,
        status="completed",
        progress=100,
        processed_units=len(segment_rows),
        total_units=len(segment_rows),
        result_ref=safe_text(job.get("id")),
        metadata={"qa_summary": summary, "result_file": report_manifest, "job_id": job.get("id")},
    )


def process_pro_task(task_id: str, payload: Dict[str, Any]) -> Dict[str, Any]:
    workspace = safe_text(payload.get("workspace") or "Unassigned")
    user_email = safe_text(payload.get("user_email") or "async-worker@cognisweep.local")
    task_payload = payload.get("payload") if isinstance(payload.get("payload"), dict) else {}
    params = task_payload.get("parameters") if isinstance(task_payload.get("parameters"), dict) else {}
    target_language = safe_text(params.get("target_language") or "Target")
    source_language = safe_text(params.get("source_language") or "English")
    domain = safe_text(params.get("domain") or "General")
    reference_context = params.get("reference_context") if isinstance(params.get("reference_context"), dict) else {}
    quality_inputs = params.get("quality_inputs") if isinstance(params.get("quality_inputs"), dict) else {}
    update_task(task_id, payload, status="running", progress=12)
    rows, rules, primary = load_task_inputs(payload)
    rows = sentence_segment_rows_for_pro(rows)
    if not rows:
        raise ValueError("No translatable rows were detected in the queued input file.")
    source_texts = [safe_text(row.get("source") or row.get("target")) for row in rows]
    protected_terms = [safe_text(item.get("term") if isinstance(item, dict) else item) for item in rules.get("dnt", []) or []]
    client_context = translation_context_metadata(rules, reference_context, source_texts, target_language, domain)
    translations: List[str] = []
    usage: Dict[str, Any] = {"provider": "none", "model": "", "success": False, "error": "No worker translation route attempted."}
    route_error = ""
    managed_mt_allowed = bool(params.get("allow_managed_amazon_translate"))
    update_task(task_id, payload, status="running", progress=35, total_units=len(rows), processed_units=0)
    if translate_batch is not None and managed_mt_allowed:
        try:
            translations, usage = translate_batch(
                source_language=source_language,
                target_language=target_language,
                texts=source_texts,
                protected_terms=protected_terms,
                metadata={
                    "workflow": "async_pro_translation",
                    "domain": domain,
                    "task_id": task_id,
                    "client_context": client_context,
                },
            )
        except Exception as exc:
            route_error = safe_text(exc)
            translations = ["" for _ in source_texts]
            usage = {"provider": "managed_mt", "model": "amazon_translate_pending", "success": False, "error": route_error, "characters": sum(len(t) for t in source_texts), "requests": 1}
    elif translate_batch is not None:
        route_error = "Managed Amazon Translate is not entitled for this workspace."
        translations = ["" for _ in source_texts]
    else:
        route_error = "translator_router is unavailable in this worker environment."
        translations = ["" for _ in source_texts]
    update_task(task_id, payload, status="running", progress=70, total_units=len(rows), processed_units=len(translations))
    review_rows: List[Dict[str, Any]] = []
    missing = 0
    for row, translated in zip(rows, translations):
        target = safe_text(translated)
        if not target or target == safe_text(row.get("source")):
            missing += 1
            status = "Needs Review"
            match = "MT unavailable"
        else:
            status = "MT"
            match = safe_text(usage.get("engine") or usage.get("provider") or "MT")
        review_row = {**row, "target": target, "status": status, "match": match}
        row_findings = qa_findings_for_row(review_row, rules)
        if row_findings:
            review_row["qa_findings"] = row_findings[:10]
            review_row["qa_summary"] = "; ".join(f"{item['Severity']}: {item['Error Category']}" for item in row_findings[:4])
            if status == "MT":
                review_row["status"] = "Needs Review"
        review_rows.append(review_row)
    summary = {
        "target_language": target_language,
        "segments": len(review_rows),
        "missing_or_review": missing,
        "status": "Needs Human Review" if missing else "Completed",
        "route": safe_text(usage.get("engine") or usage.get("provider") or "managed_mt"),
        "route_error": route_error,
    }
    export_source: Dict[str, Any] = {}
    try:
        export_source = build_export_source_asset_from_bytes(
            manifest_bytes(primary),
            manifest_file_name(primary),
            safe_text(primary.get("mime_type")),
        )
    except Exception as exc:
        LOGGER.warning("Unable to attach original source asset for Pro reconstruction export: %s", exc)

    review_job_id = save_persistent_editor_job(
        "cat",
        review_rows,
        metadata={
            "workspace": workspace,
            "user_email": user_email,
            "file_name": manifest_file_name(primary),
            "target_language": target_language,
            "source": "CogniSweep Async Pro",
            "status": "needs_review" if missing else "draft",
            "task_id": task_id,
            "rules": rules,
            "context": reference_context,
            "export_source": export_source,
            "quality_inputs": quality_inputs,
        },
        user={"email": user_email, "workspace": workspace},
    )
    workbook = create_review_workbook(review_rows, summary)
    review_manifest = persist_result_file(
        workbook,
        "CogniSweep_Pro_Human_Review.xlsx",
        workspace,
        user_email,
        "pro_review_workbook",
        task_id,
        "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
    )
    delivery_zip = create_translation_delivery_zip(review_rows, primary, review_workbook=workbook)
    delivery_manifest = persist_result_file(
        delivery_zip,
        "CogniSweep_Translation_Delivery.zip",
        workspace,
        user_email,
        "pro_translation_delivery",
        task_id,
        "application/zip",
    )
    log_persistent_usage_event(
        {**usage, "segments": len(review_rows), "characters": sum(len(t) for t in source_texts), "requests": usage.get("requests", 1)},
        purpose="pro_translation",
        segment_count=len(review_rows),
        user={"email": user_email, "workspace": workspace},
        metadata={"task_id": task_id, "review_job_id": review_job_id, "result_file_id": delivery_manifest.get("id"), "review_workbook_file_id": review_manifest.get("id")},
    )
    status = "needs_review" if missing else "completed"
    return update_task(
        task_id,
        payload,
        status=status,
        progress=100,
        processed_units=len(review_rows),
        total_units=len(review_rows),
        result_ref=review_job_id,
        metadata={
            "pro_summary": summary,
            "result_file": delivery_manifest,
            "result_files": [review_manifest],
            "review_job_id": review_job_id,
            "editor_job_id": review_job_id,
            "workflow": "pro_translation",
            "source": "CogniSweep Async Pro",
            "file_name": manifest_file_name(primary),
            "target_language": target_language,
            "quality_inputs": quality_inputs,
        },
    )


def process_task_payload(payload: Dict[str, Any]) -> Dict[str, Any]:
    task_id = task_id_from_payload(payload)
    task_payload = payload.get("payload") if isinstance(payload.get("payload"), dict) else {}
    workflow = safe_text(task_payload.get("workflow") or payload.get("task_type") or payload.get("workflow")).lower()
    try:
        if workflow in {"qa", "qa_workflow"} or "qa" in workflow:
            return process_qa_task(task_id, payload)
        if workflow in {"pro_translation", "pro", "translation"} or "pro" in workflow:
            return process_pro_task(task_id, payload)
        raise ValueError(f"Unsupported async workflow: {workflow or 'unknown'}")
    except Exception as exc:
        LOGGER.exception("Async workflow processing failed for %s", task_id)
        return update_task(task_id, payload, status="failed", progress=100, error=safe_text(exc))


def process_task_id(task_id: str) -> Dict[str, Any]:
    spool = read_spool(task_id)
    if not spool:
        raise FileNotFoundError(f"No worker spool found for task {task_id}.")
    payload = spool.get("payload") if isinstance(spool.get("payload"), dict) else {"task_id": task_id}
    payload.setdefault("task_id", task_id)
    return process_task_payload(payload)


def queued_spool_documents() -> List[Dict[str, Any]]:
    documents = []
    for path in sorted(worker_root().glob("*.json"), key=lambda item: item.stat().st_mtime):
        try:
            doc = json.loads(path.read_text(encoding="utf-8"))
        except Exception as exc:
            LOGGER.warning("Unable to read spool document %s: %s", path, exc)
            continue
        status_record = doc.get("status_record") if isinstance(doc.get("status_record"), dict) else {}
        if safe_text(status_record.get("status")).lower() in {"queued", "running", ""}:
            documents.append(doc)
    return documents


def process_next_queued_task() -> Optional[Dict[str, Any]]:
    docs = queued_spool_documents()
    if not docs:
        return None
    payload = docs[0].get("payload") if isinstance(docs[0].get("payload"), dict) else {}
    return process_task_payload(payload)


def smoke_check() -> Dict[str, Any]:
    sample = pd.DataFrame([
        {"Source": "Hello {{name}} 😊", "Target": "Bonjour {{name}}"},
        {"Source": "Save 25 files", "Target": "Enregistrer les fichiers"},
    ])
    buffer = io.BytesIO()
    sample.to_excel(buffer, index=False)
    task_id = f"processor-smoke-{uuid.uuid4().hex[:10]}"
    input_path = worker_root() / f"{task_id}.xlsx"
    input_path.write_bytes(buffer.getvalue())
    payload = {
        "task_id": task_id,
        "task_type": "qa",
        "label": "Async workflow processor smoke test",
        "workspace": "Platform",
        "user_email": "async-worker@cognisweep.local",
        "metadata": {"workflow": "qa"},
        "payload": {
            "workflow": "qa",
            "input_files": [{
                "file_name": input_path.name,
                "storage_provider": "local",
                "storage_key": str(input_path),
                "mime_type": mimetypes.guess_type(str(input_path))[0] or "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
            }],
            "rules_files": [],
            "parameters": {"strictness": "Standard", "domain": "General"},
        },
    }
    record = update_task(task_id, payload, status="queued", progress=2)
    write_spool_status(task_id, payload, record)
    final = process_task_payload(payload)
    return {"task_id": task_id, "final_status": final.get("status"), "progress": final.get("progress"), "metadata": final.get("metadata_json", {})}


def main() -> int:
    parser = argparse.ArgumentParser(description="Process queued CogniSweep async QA/Pro tasks.")
    parser.add_argument("--task-id", default="", help="Process one specific worker-spooled task id.")
    parser.add_argument("--once", action="store_true", help="Process the oldest queued worker-spooled task and exit.")
    parser.add_argument("--loop", action="store_true", help="Continuously process queued worker-spooled tasks.")
    parser.add_argument("--interval", type=int, default=int(env_value("ERRORSWEEP_ASYNC_PROCESSOR_INTERVAL", "10")))
    parser.add_argument("--smoke", action="store_true", help="Run a local deterministic processor smoke test and exit.")
    args = parser.parse_args()
    logging.basicConfig(level=safe_text(env_value("ERRORSWEEP_ASYNC_PROCESSOR_LOG_LEVEL")) or "INFO", format="%(asctime)s %(levelname)s %(message)s")
    if args.smoke:
        result = smoke_check()
        print(json.dumps(result, ensure_ascii=False, indent=2, default=safe_text))
        return 0 if result.get("final_status") == "completed" else 1
    if args.task_id:
        result = process_task_id(args.task_id)
        print(json.dumps({"task_id": args.task_id, "status": result.get("status"), "progress": result.get("progress")}, ensure_ascii=False, indent=2))
        return 0 if result.get("status") not in {"failed", ""} else 1
    if args.once:
        result = process_next_queued_task()
        print(json.dumps(result or {"processed": False, "message": "No queued worker-spooled tasks."}, ensure_ascii=False, indent=2, default=safe_text))
        return 0
    if args.loop:
        while True:
            result = process_next_queued_task()
            if result:
                LOGGER.info("Processed async task %s -> %s", result.get("id"), result.get("status"))
            time.sleep(max(1, int(args.interval or 10)))
    parser.print_help()
    return 0


if __name__ == "__main__":
    raise SystemExit(main())
