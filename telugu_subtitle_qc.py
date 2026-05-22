"""ErrorSweep Telugu Subtitle QC Engine — Phase 1/2 MVP.

This module is intentionally self-contained and API-free.
It parses Telugu subtitle/script files and returns explainable QC findings.

Supported input now:
- .srt
- .vtt
- .txt / generic event text

Main entry points:
- run_telugu_subtitle_qc(file_bytes, file_name, options=None)
- build_telugu_qc_excel(result)
"""
from __future__ import annotations

import io
import re
import unicodedata
from dataclasses import dataclass, asdict
from typing import Any, Dict, Iterable, List, Optional, Tuple

import pandas as pd
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter


TELUGU_RANGE_RE = re.compile(r"[\u0C00-\u0C7F]")
LATIN_WORD_RE = re.compile(r"\b[A-Za-z][A-Za-z0-9_\-]*\b")
TIME_SRT_RE = re.compile(
    r"(?P<start>\d{1,2}:\d{2}:\d{2}[,.]\d{1,3})\s*-->\s*(?P<end>\d{1,2}:\d{2}:\d{2}[,.]\d{1,3})(?:\s+.*)?"
)
TIME_COLON_RE = re.compile(
    r"(?P<start>\d{1,2}:\d{2}:\d{2}(?::\d{1,2})?)\s*(?:-->|-|to)\s*(?P<end>\d{1,2}:\d{2}:\d{2}(?::\d{1,2})?)"
)
EVENT_RE = re.compile(r"^\s*(?:event\s*#?|subtitle\s*#?|cue\s*#?)\s*(\d+)\s*$", re.I)
BRACKET_RE = re.compile(r"\[[^\]]*\]|\([^)]*\)")
PLACEHOLDER_RE = re.compile(r"(\{\{[^{}]+\}\}|\{[^{}]+\}|%\$?\d*[sd]|%[sd]|<[^>]+>|\$\w+)")
NUMBER_RE = re.compile(r"\d+(?:[.,:]\d+)*")

# Telugu Unicode classes.
TELUGU_CONSONANT = set(chr(cp) for cp in range(0x0C15, 0x0C3A))
TELUGU_MATRA = set(chr(cp) for cp in range(0x0C3E, 0x0C4D))
TELUGU_HALANT = "\u0C4D"
TELUGU_ANUSVARA_VISARGA = {"\u0C02", "\u0C03"}
TELUGU_DIGITS = set(chr(cp) for cp in range(0x0C66, 0x0C70))
LATIN_DIGITS = set("0123456789")

# Conservative known error pairs from the Telugu QC blueprint + production UI patterns.
KNOWN_ERROR_PAIRS: Dict[str, Tuple[str, str, str]] = {
    "చూసాను": ("చూశాను", "SP-001", "Common past-tense spelling issue: శ is expected in this form."),
    "చేసాడు": ("చేశాడు", "SP-002", "Common past-tense spelling issue: చేశాడు is the preferred form."),
    "చేసారు": ("చేశారు", "SP-003", "Common past-tense spelling issue: చేశారు is the preferred form."),
    "బహుశ": ("బహుశా", "SP-004", "Adverb usually requires final long ఆ."),
    "కొత్తేమి": ("కొత్తేమీ", "SP-005", "Negation/emphasis particle usually takes long ఈ."),
    "తేనే": ("తేనె", "SP-006", "Wrong vowel sign for the word honey."),
    "ముష్రూమ్స్": ("మష్రూమ్స్", "SP-007", "Loanword spelling looks inconsistent with expected transliteration."),
    "గుర్తిచిన": ("గుర్తించిన", "SP-008", "Missing conjunct/nasal component in the verb form."),
    "chupinchandi": ("చూపించండి", "MS-001", "Romanized Telugu appears inside Telugu output."),
    "cheyandi": ("చేయండి", "MS-002", "Romanized Telugu appears inside Telugu output."),
}

ALLOWED_LATIN_TERMS = {
    "api", "ui", "ux", "url", "http", "https", "html", "css", "json", "xml",
    "pdf", "docx", "xlsx", "csv", "srt", "vtt", "ok", "id", "otp", "sms",
    "app", "email", "login", "logout", "fitjourney", "docflow",
}

PRONOUN_FEATURES = {
    "నేను": {"person": "1", "number": "sg", "gender": "any", "allowed_endings": ["ాను", "తున్నాను"]},
    "నువ్వు": {"person": "2", "number": "sg", "gender": "any", "allowed_endings": ["ావు", "తున్నావు"]},
    "నీవు": {"person": "2", "number": "sg", "gender": "any", "allowed_endings": ["ావు", "తున్నావు"]},
    "అతను": {"person": "3", "number": "sg", "gender": "m", "allowed_endings": ["ాడు", "తున్నాడు"]},
    "వాడు": {"person": "3", "number": "sg", "gender": "m", "allowed_endings": ["ాడు", "తున్నాడు"]},
    "ఆమె": {"person": "3", "number": "sg", "gender": "f", "allowed_endings": ["ింది", "తోంది", "తున్నది"]},
    "ఆవిడ": {"person": "3", "number": "sg", "gender": "f", "allowed_endings": ["ింది", "తోంది", "తున్నది"]},
    "మనం": {"person": "1", "number": "pl", "gender": "any", "allowed_endings": ["ాం", "తున్నాం"]},
    "మేము": {"person": "1", "number": "pl", "gender": "any", "allowed_endings": ["ాం", "తున్నాం"]},
    "మీరు": {"person": "2", "number": "pl", "gender": "any", "allowed_endings": ["ారు", "తున్నారు"]},
    "వాళ్ళు": {"person": "3", "number": "pl", "gender": "human", "allowed_endings": ["ారు", "తున్నారు"]},
    "వారు": {"person": "3", "number": "pl", "gender": "human", "allowed_endings": ["ారు", "తున్నారు"]},
    "అవి": {"person": "3", "number": "pl", "gender": "inanimate", "allowed_endings": ["ాయి", "తున్నాయి"]},
}

ONGOING_STAGE_TERMS = ["మోగ", "శబ్ద", "అలారం", "తలుపు", "కారు", "ఇంజిన్", "వర్ష", "పడ", "నడుస్త", "వస్త"]

@dataclass
class SubtitleEvent:
    number: int
    start_ms: Optional[int]
    end_ms: Optional[int]
    raw_text: str
    dialogue: str
    stage_directions: List[str]
    speakers: List[str]
    is_dual_speaker: bool
    is_omitted: bool

    @property
    def duration_ms(self) -> Optional[int]:
        if self.start_ms is None or self.end_ms is None:
            return None
        return self.end_ms - self.start_ms

    @property
    def char_count(self) -> int:
        return len(self.dialogue)

    @property
    def timecode(self) -> str:
        if self.start_ms is None or self.end_ms is None:
            return ""
        return f"{ms_to_time(self.start_ms)} --> {ms_to_time(self.end_ms)}"

    def to_dict(self) -> Dict[str, Any]:
        d = asdict(self)
        d["duration_ms"] = self.duration_ms
        d["char_count"] = self.char_count
        d["timecode"] = self.timecode
        return d

@dataclass
class QCFinding:
    event_number: int
    timecode: str
    layer: str
    severity: str
    rule_id: str
    description: str
    original: str
    suggestion: str
    context: str
    confidence: float

    def to_dict(self) -> Dict[str, Any]:
        return asdict(self)


def normalize_text(text: Any) -> str:
    text = "" if text is None else str(text)
    text = unicodedata.normalize("NFC", text)
    text = text.replace("\u200b", "").replace("\u200d", "")
    text = text.replace("\u00a0", " ")
    return text.strip()


def decode_bytes(data: bytes) -> Tuple[str, str]:
    for enc in ("utf-8-sig", "utf-8", "cp1252", "latin-1"):
        try:
            return data.decode(enc), enc
        except UnicodeDecodeError:
            pass
    return data.decode("utf-8", errors="replace"), "utf-8-replace"


def time_to_ms(t: str, fps: int = 25) -> int:
    t = t.strip().replace(",", ".")
    parts = t.split(":")
    if len(parts) == 4:  # HH:MM:SS:FF
        h, m, s, f = [int(x) for x in parts]
        return ((h * 3600 + m * 60 + s) * 1000) + int((f / max(fps, 1)) * 1000)
    h, m, rest = parts
    if "." in rest:
        s, ms = rest.split(".", 1)
        ms = int((ms + "000")[:3])
    else:
        s, ms = rest, 0
    return (int(h) * 3600 + int(m) * 60 + int(s)) * 1000 + ms


def ms_to_time(ms: int) -> str:
    if ms is None:
        return ""
    h = ms // 3600000
    ms %= 3600000
    m = ms // 60000
    ms %= 60000
    s = ms // 1000
    ms %= 1000
    return f"{h:02d}:{m:02d}:{s:02d},{ms:03d}"


def extract_stage_directions(text: str) -> List[str]:
    return [m.group(0) for m in BRACKET_RE.finditer(text or "")]


def strip_stage_directions(text: str) -> str:
    return normalize_text(BRACKET_RE.sub(" ", text or ""))


def extract_speakers(lines: List[str]) -> List[str]:
    speakers = []
    for line in lines:
        m = re.match(r"^\s*\[([^\]]{1,40})\]\s*[:：-]?", line)
        if m:
            speakers.append(m.group(1).strip())
        m2 = re.match(r"^\s*([A-Za-z\u0C00-\u0C7F ]{2,40})\s*[:：]\s+", line)
        if m2:
            speakers.append(m2.group(1).strip())
    return list(dict.fromkeys(speakers))


def make_event(number: int, start_ms: Optional[int], end_ms: Optional[int], text_lines: List[str]) -> SubtitleEvent:
    raw = "\n".join(line.rstrip() for line in text_lines).strip()
    stage = extract_stage_directions(raw)
    dialogue = strip_stage_directions(raw)
    dialogue = re.sub(r"\s+", " ", dialogue).strip()
    is_dual = sum(1 for line in text_lines if line.strip().startswith("-")) >= 2
    is_omitted = bool(re.search(r"\bOMITTED\b|\[\s*omitted\s*\]", raw, re.I))
    return SubtitleEvent(
        number=number,
        start_ms=start_ms,
        end_ms=end_ms,
        raw_text=raw,
        dialogue=dialogue,
        stage_directions=stage,
        speakers=extract_speakers(text_lines),
        is_dual_speaker=is_dual,
        is_omitted=is_omitted,
    )


def parse_srt(text: str, fps: int = 25) -> List[SubtitleEvent]:
    text = text.replace("\r\n", "\n").replace("\r", "\n")
    blocks = re.split(r"\n\s*\n", text.strip())
    events: List[SubtitleEvent] = []
    fallback_number = 1
    for block in blocks:
        lines = [line for line in block.split("\n") if line.strip()]
        if not lines:
            continue
        number = fallback_number
        if lines[0].strip().isdigit():
            number = int(lines.pop(0).strip())
        time_line_idx = None
        start_ms = end_ms = None
        for idx, line in enumerate(lines[:3]):
            m = TIME_SRT_RE.search(line) or TIME_COLON_RE.search(line)
            if m:
                start_ms = time_to_ms(m.group("start"), fps=fps)
                end_ms = time_to_ms(m.group("end"), fps=fps)
                time_line_idx = idx
                break
        if time_line_idx is None:
            continue
        text_lines = lines[time_line_idx + 1:]
        events.append(make_event(number, start_ms, end_ms, text_lines))
        fallback_number = number + 1
    return events


def parse_vtt(text: str, fps: int = 25) -> List[SubtitleEvent]:
    text = re.sub(r"^WEBVTT.*?\n", "", text.replace("\r\n", "\n").replace("\r", "\n"), flags=re.I | re.S)
    return parse_srt(text, fps=fps)


def parse_event_text(text: str, fps: int = 25) -> List[SubtitleEvent]:
    lines = text.replace("\r\n", "\n").replace("\r", "\n").split("\n")
    events: List[SubtitleEvent] = []
    current_num: Optional[int] = None
    current_lines: List[str] = []
    current_start = current_end = None

    def flush():
        nonlocal current_num, current_lines, current_start, current_end
        if current_lines:
            num = current_num or len(events) + 1
            events.append(make_event(num, current_start, current_end, current_lines))
        current_num = None
        current_lines = []
        current_start = current_end = None

    for line in lines:
        em = EVENT_RE.match(line)
        if em:
            flush()
            current_num = int(em.group(1))
            continue
        tm = TIME_SRT_RE.search(line) or TIME_COLON_RE.search(line)
        if tm:
            current_start = time_to_ms(tm.group("start"), fps=fps)
            current_end = time_to_ms(tm.group("end"), fps=fps)
            continue
        if line.strip():
            current_lines.append(line)
        else:
            # Keep simple TXT as one event per non-empty line if there is no explicit event marker.
            if current_num is None and current_lines:
                flush()
    flush()
    return events


def parse_subtitle_events(file_bytes: bytes, file_name: str, fps: int = 25) -> Tuple[List[SubtitleEvent], Dict[str, Any]]:
    text, encoding = decode_bytes(file_bytes)
    lower = file_name.lower()
    if lower.endswith(".srt"):
        events = parse_srt(text, fps=fps)
        fmt = "SRT"
    elif lower.endswith(".vtt"):
        events = parse_vtt(text, fps=fps)
        fmt = "VTT"
    else:
        # Try SRT-style first; fall back to generic text.
        events = parse_srt(text, fps=fps)
        fmt = "SRT-like"
        if not events:
            events = parse_event_text(text, fps=fps)
            fmt = "Text"
    meta = {"file_name": file_name, "encoding": encoding, "format": fmt, "event_count": len(events)}
    return events, meta


def finding(event: SubtitleEvent, layer: str, severity: str, rule_id: str, description: str, original: str, suggestion: str = "", confidence: float = 1.0) -> QCFinding:
    return QCFinding(
        event_number=event.number,
        timecode=event.timecode,
        layer=layer,
        severity=severity,
        rule_id=rule_id,
        description=description,
        original=original,
        suggestion=suggestion,
        context=event.raw_text[:700],
        confidence=round(float(confidence), 3),
    )


def contains_telugu(text: str) -> bool:
    return bool(TELUGU_RANGE_RE.search(text or ""))


def validate_unicode(event: SubtitleEvent) -> List[QCFinding]:
    out: List[QCFinding] = []
    text = event.raw_text
    chars = list(text)
    for i, ch in enumerate(chars):
        prev_ch = chars[i - 1] if i > 0 else ""
        next_ch = chars[i + 1] if i + 1 < len(chars) else ""
        if ch in TELUGU_MATRA:
            if not prev_ch or prev_ch.isspace() or prev_ch.isascii() or prev_ch in TELUGU_MATRA:
                out.append(finding(event, "Unicode", "Critical", "UN-001", "Telugu dependent vowel sign/matra appears in an invalid position.", ch, "Check surrounding Telugu syllable.", 0.98))
        if ch == TELUGU_HALANT:
            if not next_ch or next_ch not in TELUGU_CONSONANT:
                out.append(finding(event, "Unicode", "Critical", "UN-002", "Telugu halant/virama is not followed by a consonant.", ch, "Remove stray halant or restore missing consonant.", 0.98))
    if any(ch in TELUGU_DIGITS for ch in text) and any(ch in LATIN_DIGITS for ch in text):
        out.append(finding(event, "Unicode", "Info", "UN-003", "Telugu and Latin digits are mixed in the same subtitle event.", event.raw_text, "Use one digit style consistently.", 0.7))
    return out


def spelling_checks(event: SubtitleEvent) -> List[QCFinding]:
    out: List[QCFinding] = []
    text = event.raw_text
    for wrong, (right, rule_id, reason) in KNOWN_ERROR_PAIRS.items():
        if wrong in text:
            out.append(finding(event, "Spelling", "Warning" if rule_id.startswith("SP") else "Major", rule_id, reason, wrong, right, 0.95))
    return out


def formatting_checks(event: SubtitleEvent) -> List[QCFinding]:
    out: List[QCFinding] = []
    t = event.raw_text
    if re.search(r" {2,}", t):
        out.append(finding(event, "Structure", "Info", "ST-001", "Multiple consecutive spaces found.", "multiple spaces", re.sub(r" {2,}", " ", t), 0.9))
    if t.count("[") != t.count("]"):
        out.append(finding(event, "Structure", "Warning", "ST-002", "Unbalanced square brackets in subtitle text.", t, "Balance or remove bracketed stage direction.", 0.95))
    if t.count("(") != t.count(")"):
        out.append(finding(event, "Structure", "Warning", "ST-003", "Unbalanced parentheses in subtitle text.", t, "Balance parentheses.", 0.9))
    if event.is_dual_speaker and "\n" not in event.raw_text:
        out.append(finding(event, "Structure", "Info", "ST-004", "Dual-speaker marker detected but text is not split into separate lines.", event.raw_text, "Place each speaker on a separate line if required by client style.", 0.6))
    return out


def timing_checks(events: List[SubtitleEvent]) -> List[QCFinding]:
    out: List[QCFinding] = []
    prev: Optional[SubtitleEvent] = None
    for ev in events:
        dur = ev.duration_ms
        if ev.start_ms is not None and ev.end_ms is not None:
            if ev.start_ms >= ev.end_ms:
                out.append(finding(ev, "Timing", "Critical", "TM-001", "Subtitle in-time is not before out-time.", ev.timecode, "Fix timecode range.", 1.0))
            elif dur is not None and dur < 800 and ev.dialogue:
                out.append(finding(ev, "Timing", "Warning", "TM-002", "Subtitle duration is very short for readable text.", ev.timecode, "Increase duration or shorten text.", 0.85))
            elif dur is not None and dur > 7000 and ev.dialogue:
                out.append(finding(ev, "Timing", "Info", "TM-003", "Subtitle duration is unusually long.", ev.timecode, "Review timing and segmentation.", 0.75))
        if prev and ev.start_ms is not None and prev.end_ms is not None and ev.start_ms < prev.end_ms:
            out.append(finding(ev, "Timing", "Warning", "TM-004", "Subtitle overlaps with the previous event.", ev.timecode, "Adjust start/end times to avoid unintended overlap.", 0.9))
        prev = ev
    return out


def mixed_script_checks(event: SubtitleEvent) -> List[QCFinding]:
    out: List[QCFinding] = []
    text = event.raw_text
    if not contains_telugu(text):
        return out
    for m in LATIN_WORD_RE.finditer(text):
        word = m.group(0)
        if word.lower() in ALLOWED_LATIN_TERMS:
            continue
        if PLACEHOLDER_RE.fullmatch(word):
            continue
        # Most random Romanized words inside Telugu subtitle should be reviewed.
        out.append(finding(event, "Mixed Script", "Warning", "MS-003", "Unexpected Latin/Roman word appears inside Telugu subtitle.", word, "Translate or transliterate consistently unless it is a protected product term.", 0.75))
    return out


def grammar_checks(event: SubtitleEvent) -> List[QCFinding]:
    out: List[QCFinding] = []
    text = event.dialogue
    if not contains_telugu(text):
        return out
    for pronoun, info in PRONOUN_FEATURES.items():
        if re.search(rf"(?<![\u0C00-\u0C7F]){re.escape(pronoun)}(?![\u0C00-\u0C7F])", text):
            endings = info["allowed_endings"]
            # Look at the final Telugu token only to avoid too many false positives.
            telugu_tokens = re.findall(r"[\u0C00-\u0C7F]+", text)
            if telugu_tokens:
                final = telugu_tokens[-1]
                if not any(final.endswith(e) for e in endings):
                    out.append(finding(event, "Grammar", "Info", "GR-001", f"Possible subject-verb agreement issue for pronoun '{pronoun}'.", final, f"Verb ending should agree with {pronoun}; expected one of: {', '.join(endings)}", 0.55))
            break
    # Present continuous vs habitual, very conservative.
    if any(term in " ".join(event.stage_directions + [event.dialogue]) for term in ONGOING_STAGE_TERMS):
        if "మోగుతుంది" in text:
            out.append(finding(event, "Grammar", "Warning", "GR-004", "Possible aspect issue: ongoing action may require continuous form.", "మోగుతుంది", "మోగుతోంది", 0.7))
    return out


def duplicate_checks(events: List[SubtitleEvent]) -> List[QCFinding]:
    out: List[QCFinding] = []
    seen: Dict[str, SubtitleEvent] = {}
    prev_text = ""
    for ev in events:
        normalized = re.sub(r"\s+", " ", ev.dialogue.strip())
        if not normalized or ev.is_omitted:
            continue
        if normalized == prev_text:
            out.append(finding(ev, "Consistency", "Warning", "CO-002", "Consecutive duplicate subtitle text found.", ev.dialogue, "Check if this repeated subtitle is intentional.", 0.85))
        if normalized in seen and ev.number != seen[normalized].number:
            # Non-consecutive duplicates can be normal. Keep as info only.
            out.append(finding(ev, "Consistency", "Info", "CO-003", f"Subtitle text also appeared earlier in event {seen[normalized].number}.", ev.dialogue, "Review only if repetition is unintended.", 0.5))
        else:
            seen[normalized] = ev
        prev_text = normalized
    return out


def consistency_checks(events: List[SubtitleEvent]) -> List[QCFinding]:
    out: List[QCFinding] = []
    # Known loanword/proper noun variants from blueprint examples.
    variant_groups = [
        ["కజకిస్తాన్", "కజకస్తాన్"],
        ["ఫిన్ల్యాండ్", "ఫిన్లాండ్"],
        ["మిస్సైల్", "మిసైల్"],
        ["చైర్మన్", "ఛైర్మన్"],
        ["వెల్కమ్ స్క్రీన్", "స్వాగత తెర"],
        ["డాష్‌బోర్డ్", "డాష్బోర్డ్"],
    ]
    text_to_event = [(ev, ev.raw_text) for ev in events]
    for group in variant_groups:
        found: Dict[str, List[SubtitleEvent]] = {}
        for term in group:
            for ev, text in text_to_event:
                if term in text:
                    found.setdefault(term, []).append(ev)
        if len(found) > 1:
            canonical = max(found.items(), key=lambda kv: len(kv[1]))[0]
            for term, evs in found.items():
                if term == canonical:
                    continue
                for ev in evs:
                    out.append(finding(ev, "Consistency", "Warning", "CO-001", "Inconsistent spelling/terminology variant found in the same file.", term, canonical, 0.8))
    out.extend(duplicate_checks(events))
    return out


def stage_direction_checks(event: SubtitleEvent) -> List[QCFinding]:
    out: List[QCFinding] = []
    for sd in event.stage_directions:
        inner = sd.strip("[]() ")
        if not inner:
            continue
        if contains_telugu(inner) and any(w in inner for w in ["మనిషి", "పురుషుడు", "అమ్మాయి", "మహిళ", "వ్యక్తి"]):
            if "మోగుతోంది" in inner:
                out.append(finding(event, "Semantic", "Info", "SE-002", "Stage direction may use inanimate verb form with a human subject.", inner, "Check gender/person agreement for stage direction.", 0.5))
        if re.search(r"\b(laughs|cries|coughs|music|alarm)\b", inner, flags=re.I):
            out.append(finding(event, "Semantic", "Info", "SE-003", "English stage direction appears in subtitle text.", sd, "Translate stage direction to Telugu if client style requires localized SDH cues.", 0.7))
    return out


def classify_verdict(findings: List[QCFinding]) -> str:
    severities = {f.severity for f in findings}
    if "Critical" in severities:
        return "Fail"
    if "Warning" in severities:
        return "Conditional Pass"
    return "Pass"


def run_telugu_subtitle_qc(file_bytes: bytes, file_name: str, options: Optional[Dict[str, Any]] = None) -> Dict[str, Any]:
    options = options or {}
    fps = int(options.get("fps", 25) or 25)
    events, meta = parse_subtitle_events(file_bytes, file_name, fps=fps)
    findings: List[QCFinding] = []

    findings.extend(timing_checks(events))
    findings.extend(consistency_checks(events))

    for ev in events:
        if ev.is_omitted:
            continue
        findings.extend(validate_unicode(ev))
        findings.extend(formatting_checks(ev))
        findings.extend(spelling_checks(ev))
        findings.extend(mixed_script_checks(ev))
        findings.extend(grammar_checks(ev))
        findings.extend(stage_direction_checks(ev))

    # Deduplicate by event + rule + original.
    dedup: Dict[Tuple[int, str, str], QCFinding] = {}
    severity_rank = {"Critical": 3, "Warning": 2, "Info": 1}
    for f in findings:
        key = (f.event_number, f.rule_id, f.original)
        existing = dedup.get(key)
        if not existing or severity_rank.get(f.severity, 0) > severity_rank.get(existing.severity, 0):
            dedup[key] = f
    findings = sorted(dedup.values(), key=lambda f: (-(severity_rank.get(f.severity, 0)), f.event_number, f.rule_id))

    summary = {
        **meta,
        "findings_count": len(findings),
        "critical_count": sum(1 for f in findings if f.severity == "Critical"),
        "warning_count": sum(1 for f in findings if f.severity == "Warning"),
        "info_count": sum(1 for f in findings if f.severity == "Info"),
        "verdict": classify_verdict(findings),
    }
    return {
        "summary": summary,
        "events": [ev.to_dict() for ev in events],
        "findings": [f.to_dict() for f in findings],
    }


def _style_sheet(ws, freeze: str = "A2"):
    header_fill = PatternFill("solid", fgColor="1F4E79")
    header_font = Font(color="FFFFFF", bold=True)
    thin = Side(style="thin", color="D9E2F3")
    border = Border(left=thin, right=thin, top=thin, bottom=thin)
    for cell in ws[1]:
        cell.fill = header_fill
        cell.font = header_font
        cell.alignment = Alignment(wrap_text=True, vertical="top")
        cell.border = border
    ws.freeze_panes = freeze
    for row in ws.iter_rows():
        for cell in row:
            cell.alignment = Alignment(wrap_text=True, vertical="top")
            cell.border = border
    for col in ws.columns:
        letter = get_column_letter(col[0].column)
        max_len = 12
        for cell in col[:200]:
            max_len = max(max_len, min(len(str(cell.value or "")), 70))
        ws.column_dimensions[letter].width = min(max_len + 2, 70)


def build_telugu_qc_excel(result: Dict[str, Any]) -> bytes:
    wb = Workbook()
    ws = wb.active
    ws.title = "Telugu QC Findings"
    headers = ["Event No", "Timecode", "Layer", "Severity", "Rule ID", "Original", "Suggestion", "Description", "Context", "Confidence"]
    ws.append(headers)
    for f in result.get("findings", []):
        ws.append([
            f.get("event_number", ""),
            f.get("timecode", ""),
            f.get("layer", ""),
            f.get("severity", ""),
            f.get("rule_id", ""),
            f.get("original", ""),
            f.get("suggestion", ""),
            f.get("description", ""),
            f.get("context", ""),
            f.get("confidence", ""),
        ])
    _style_sheet(ws)

    ws2 = wb.create_sheet("Summary")
    ws2.append(["Metric", "Value"])
    for k, v in result.get("summary", {}).items():
        ws2.append([k, v])
    _style_sheet(ws2)

    ws3 = wb.create_sheet("Rule Reference")
    ws3.append(["Layer", "Severity", "Rule ID", "Description"])
    reference_rows = [
        ("Timing", "Critical/Warning/Info", "TM-*", "Checks invalid, overlapping, too short, or unusually long subtitle timing."),
        ("Unicode", "Critical", "UN-*", "Checks Telugu matra/halant integrity and mixed digit styles."),
        ("Spelling", "Warning", "SP-*", "Known Telugu error pairs and spelling variants."),
        ("Mixed Script", "Warning", "MS-*", "Flags unexpected Roman/Latin words inside Telugu output."),
        ("Grammar", "Info/Warning", "GR-*", "Conservative pronoun/verb agreement and aspect checks."),
        ("Consistency", "Info/Warning", "CO-*", "Flags duplicate subtitles and known terminology variants."),
        ("Semantic", "Info", "SE-*", "Flags stage direction localization and context-review items."),
    ]
    for row in reference_rows:
        ws3.append(row)
    _style_sheet(ws3)

    ws4 = wb.create_sheet("Parsed Events")
    event_headers = ["Event No", "Timecode", "Duration ms", "Char Count", "Dialogue", "Stage Directions", "Raw Text"]
    ws4.append(event_headers)
    for ev in result.get("events", []):
        ws4.append([
            ev.get("number", ""), ev.get("timecode", ""), ev.get("duration_ms", ""), ev.get("char_count", ""),
            ev.get("dialogue", ""), "; ".join(ev.get("stage_directions", []) or []), ev.get("raw_text", ""),
        ])
    _style_sheet(ws4)

    bio = io.BytesIO()
    wb.save(bio)
    return bio.getvalue()


def findings_dataframe(result: Dict[str, Any]) -> pd.DataFrame:
    findings = result.get("findings", [])
    if not findings:
        return pd.DataFrame(columns=["event_number", "timecode", "layer", "severity", "rule_id", "original", "suggestion", "description", "confidence"])
    return pd.DataFrame(findings)


def events_dataframe(result: Dict[str, Any]) -> pd.DataFrame:
    events = result.get("events", [])
    if not events:
        return pd.DataFrame(columns=["number", "timecode", "dialogue"])
    return pd.DataFrame(events)


